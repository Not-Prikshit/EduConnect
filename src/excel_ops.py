# src/excel_ops.py
import os
from openpyxl import Workbook, load_workbook

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
DATA_DIR = os.path.abspath(DATA_DIR)
FACULTY_FILE = os.path.join(DATA_DIR, "faculty_data.xlsx")
STUDENT_FILE = os.path.join(DATA_DIR, "student_data.xlsx")

FACULTY_HEADERS = [
    "Faculty Name", "Department", "Date", "Start Time", "End Time",
    "Status", "Student Name", "Student Email", "Faculty Email", "Slot ID"
]

STUDENT_HEADERS = [
    "Student Name", "Student Email", "Faculty Name", "Department",
    "Date", "Start Time", "End Time", "Status", "Slot ID"
]

def ensure_files_exist():
    os.makedirs(DATA_DIR, exist_ok=True)
    if not os.path.exists(FACULTY_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(FACULTY_HEADERS)
        wb.save(FACULTY_FILE)
    if not os.path.exists(STUDENT_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(STUDENT_HEADERS)
        wb.save(STUDENT_FILE)

def _load_wb(path):
    return load_workbook(path)

def _save_wb(wb, path):
    tmp = path + ".tmp"
    wb.save(tmp)
    os.replace(tmp, path)

def _make_slot_id(faculty_name, date, start_time):
    fn = "".join(x for x in str(faculty_name) if x.isalnum())[:12]
    return f"{fn}_{str(date).replace('-','')}_{str(start_time).replace(':','')}"

def get_free_slots():
    ensure_files_exist()
    wb = _load_wb(FACULTY_FILE)
    ws = wb.active
    slots = []
    for row in list(ws.iter_rows(min_row=2, values_only=True)):
        status = (row[5] or "").strip()
        if status.lower() == "free":
            slot_id = row[9] or _make_slot_id(row[0], row[2], row[3])
            slots.append({
                "faculty_name": row[0],
                "department": row[1],
                "date": row[2],
                "start_time": row[3],
                "end_time": row[4],
                "status": status,
                "slot_id": slot_id,
                "faculty_email": row[8] or ""
            })
    wb.close()
    return slots

def book_slot(faculty, date, start_time, student_name, student_email):
    ensure_files_exist()
    wb = _load_wb(FACULTY_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        fname = row[0].value
        rdate = row[2].value
        rstart = row[3].value
        status = (row[5].value or "").strip()
        if fname == faculty and rdate == date and rstart == start_time and status.lower() == "free":
            row[5].value = "Requested"
            row[6].value = student_name
            row[7].value = student_email
            if not row[9].value:
                row[9].value = _make_slot_id(fname, rdate, rstart)
            _save_wb(wb, FACULTY_FILE)
            wb.close()
            _append_student_booking(student_name, student_email, fname, row[1].value, date, start_time, row[4].value, "Requested", row[9].value)
            return True, "Slot requested successfully!"
    wb.close()
    return False, "No matching free slot found."

def add_faculty_slot(faculty_name, department, date, start_time, end_time, faculty_email=""):
    ensure_files_exist()
    wb = _load_wb(FACULTY_FILE)
    ws = wb.active
    slot_id = _make_slot_id(faculty_name, date, start_time)
    ws.append([faculty_name, department, date, start_time, end_time, "Free", "", "", faculty_email, slot_id])
    _save_wb(wb, FACULTY_FILE)
    wb.close()
    return True, "Slot added successfully!"

def get_requests():
    ensure_files_exist()
    wb = _load_wb(FACULTY_FILE)
    ws = wb.active
    reqs = []
    for row in list(ws.iter_rows(min_row=2, values_only=True)):
        status = (row[5] or "").strip()
        if status.lower() == "requested":
            slot_id = row[9] or _make_slot_id(row[0], row[2], row[3])
            reqs.append({
                "faculty_name": row[0],
                "department": row[1],
                "date": row[2],
                "start_time": row[3],
                "end_time": row[4],
                "status": status,
                "student_name": row[6],
                "student_email": row[7],
                "faculty_email": row[8] or "",
                "slot_id": slot_id
            })
    wb.close()
    return reqs

def approve_request(slot_id):
    ensure_files_exist()
    wb = _load_wb(FACULTY_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        sid = (row[9].value or _make_slot_id(row[0].value, row[2].value, row[3].value))
        if sid == slot_id and (row[5].value or "").strip().lower() == "requested":
            row[5].value = "Booked"
            slot = {
                "faculty_name": row[0].value,
                "department": row[1].value,
                "date": row[2].value,
                "start_time": row[3].value,
                "end_time": row[4].value,
                "status": row[5].value,
                "student_name": row[6].value,
                "student_email": row[7].value,
                "faculty_email": row[8].value,
                "slot_id": sid
            }
            _save_wb(wb, FACULTY_FILE)
            wb.close()
            _update_student_status(slot["student_email"], slot_id, "Booked")
            return slot
    wb.close()
    return None

def _append_student_booking(student_name, student_email, faculty_name, department, date, start, end, status, slot_id):
    ensure_files_exist()
    wb = _load_wb(STUDENT_FILE)
    ws = wb.active
    ws.append([student_name, student_email, faculty_name, department, date, start, end, status, slot_id])
    _save_wb(wb, STUDENT_FILE)
    wb.close()

def _update_student_status(student_email, slot_id, new_status):
    ensure_files_exist()
    wb = _load_wb(STUDENT_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if (row[1].value == student_email) and (row[8].value == slot_id):
            row[7].value = new_status
            _save_wb(wb, STUDENT_FILE)
            wb.close()
            return True
    wb.close()
    return False

# ---------- added helpers ----------
def get_all_faculty_names():
    ensure_files_exist()
    wb = _load_wb(FACULTY_FILE)
    ws = wb.active
    names = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            names.append(row[0])
    wb.close()
    seen = set()
    uniq = []
    for n in names:
        if n not in seen:
            seen.add(n)
            uniq.append(n)
    return uniq

def get_slots_by_faculty(faculty_name):
    ensure_files_exist()
    wb = _load_wb(FACULTY_FILE)
    ws = wb.active
    slots = []
    for row in list(ws.iter_rows(min_row=2, values_only=True)):
        if (row[0] or "") == faculty_name:
            slot_id = row[9] or _make_slot_id(row[0], row[2], row[3])
            slots.append({
                "faculty_name": row[0],
                "department": row[1],
                "date": row[2],
                "start_time": row[3],
                "end_time": row[4],
                "status": row[5],
                "student_name": row[6],
                "student_email": row[7],
                "faculty_email": row[8],
                "slot_id": slot_id
            })
    wb.close()
    return slots

def faculty_exists(faculty_name):
    ensure_files_exist()
    wb = _load_wb(FACULTY_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if (row[0] or "") == faculty_name:
            wb.close()
            return True
    wb.close()
    return False

def get_student_pending(student_email):
    ensure_files_exist()
    wb = _load_wb(STUDENT_FILE)
    ws = wb.active
    res = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if (row[1] or "") == student_email:
            res.append({
                "student_name": row[0],
                "student_email": row[1],
                "faculty_name": row[2],
                "department": row[3],
                "date": row[4],
                "start_time": row[5],
                "end_time": row[6],
                "status": row[7],
                "slot_id": row[8]
            })
    wb.close()
    return res

def reject_request(slot_id):
    ensure_files_exist()
    wb = _load_wb(FACULTY_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        sid = (row[9].value or _make_slot_id(row[0].value, row[2].value, row[3].value))
        if sid == slot_id and (row[5].value or "").strip().lower() == "requested":
            row[5].value = "Free"
            row[6].value = ""
            row[7].value = ""
            _save_wb(wb, FACULTY_FILE)
            wb.close()
            _update_student_status_by_slot(slot_id, "Rejected")
            return True
    wb.close()
    return False

def _update_student_status_by_slot(slot_id, new_status):
    ensure_files_exist()
    wb = _load_wb(STUDENT_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if (row[8].value == slot_id):
            row[7].value = new_status
            _save_wb(wb, STUDENT_FILE)
            wb.close()
            return True
    wb.close()
    return False
